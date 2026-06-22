"""Prepara el Excel de comercial para subir ordenes de venta a Dynamics."""
from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE_DIR = Path(__file__).parent
SALIDA = BASE_DIR / "Layout Comercial - Preparado.xlsx"

CAPTURA_HEADERS = [
    "Cliente",
    "Código",
    "Producto",
    "Piezas",
    "Precio",
    "Órden de compra",
    "Fecha de entrega",
    "OrdenVenta",
]

RESULTADO_HEADERS = [
    "Estado",
    "Pedido Dynamics",
    "Fecha de ejecución",
    "Usuario",
    "Error",
]

HISTORIAL_HEADERS = [
    "FechaHora",
    "OrdenVenta",
    "ReferenciaCliente",
    "Cliente",
    "DescripcionPedido",
    "Codigo_Articulo",
    "Cantidad",
    "PrecioUnitario",
    "Porcentaje de descuento",
    "Fecha de envio",
    "Comentario",
]

INSTRUCCIONES = [
    ("A1", "SUBIR ORDENES DE VENTA A DYNAMICS (Comercial)"),
    ("A3", "1. Arranque dynamics-integration (run.ps1) y ngrok http 8080"),
    ("A4", "2. Pegue la URL https de ngrok en Resultado!B1 (misma que marketing)"),
    ("A5", "3. Botones: ProbarConexion, CrearCabeceraComercial, CrearLineasComercial"),
    ("A7", "HOJA CAPTURA - una orden por archivo:"),
    ("A8", "- Cliente = codigo Dynamics (ej. C0010) en todas las filas"),
    ("A9", "- Una fila por producto: Codigo, Piezas, Precio"),
    ("A10", "- Orden de compra y Fecha de entrega obligatorias"),
    ("A11", "- OrdenVenta se llena al crear cabecera; no editar"),
    ("A13", "FLUJO: 1) Llene filas  2) Generar cabecera  3) Generar lineas"),
    ("A14", "Al terminar lineas: historial + tabla Captura vacia."),
]

FILA_PRUEBA = {
    "Cliente": "C0010",
    "Código": "503635001090",
    "Producto": "ISOFLAVONAS 90 TAB",
    "Piezas": 1,
    "Precio": 515.51,
    "Órden de compra": "OC-PRUEBA-001",
    "Fecha de entrega": "2026-06-20",
    "OrdenVenta": "",
}

TEST_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
TEST_FONT = Font(bold=True)


def quitar_hoja_si_existe(wb, nombre: str) -> None:
    if nombre in wb.sheetnames:
        del wb[nombre]


def crear_hoja_instrucciones(wb) -> None:
    quitar_hoja_si_existe(wb, "Instrucciones Dynamics")
    ws = wb.create_sheet("Instrucciones Dynamics", 0)
    for addr, texto in INSTRUCCIONES:
        ws[addr] = texto
    ws.column_dimensions["A"].width = 95


def crear_hoja_resultado(wb) -> None:
    quitar_hoja_si_existe(wb, "Resultado")
    ws = wb.create_sheet("Resultado", 0)
    ws["A1"] = "URL API (ngrok)"
    ws["B1"] = ""
    ws["A2"] = "Resultado ultima ejecucion"
    for col, header in enumerate(RESULTADO_HEADERS, start=1):
        ws.cell(3, col, header)
    ws.cell(4, 1, "")

    tabla = Table(displayName="tblResultados", ref="A3:E4")
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["E"].width = 50


def crear_hoja_historial(wb) -> None:
    quitar_hoja_si_existe(wb, "Historial")
    ws = wb.create_sheet("Historial")
    for col, header in enumerate(HISTORIAL_HEADERS, start=1):
        ws.cell(1, col, header)
    # Excel exige al menos una fila de datos ademas del encabezado.
    for col in range(1, len(HISTORIAL_HEADERS) + 1):
        ws.cell(2, col, "")

    last_col = get_column_letter(len(HISTORIAL_HEADERS))
    tabla = Table(displayName="tbl_historial", ref=f"A1:{last_col}2")
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)


def crear_hoja_captura(wb) -> None:
    quitar_hoja_si_existe(wb, "Captura")
    ws = wb.create_sheet("Captura")

    for col, header in enumerate(CAPTURA_HEADERS, start=1):
        ws.cell(1, col, header)

    last_col = get_column_letter(len(CAPTURA_HEADERS))
    tabla = Table(displayName="tblCapturaComercial", ref=f"A1:{last_col}2")
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)

    anchos = {
        "A": 12,
        "B": 16,
        "C": 28,
        "D": 10,
        "E": 12,
        "F": 18,
        "G": 18,
        "H": 14,
    }
    for col, width in anchos.items():
        ws.column_dimensions[col].width = width

    for col, header in enumerate(CAPTURA_HEADERS, start=1):
        valor = FILA_PRUEBA.get(header, "")
        celda = ws.cell(2, col, valor)
        celda.fill = copy(TEST_FILL)
        celda.font = copy(TEST_FONT)


def preparar_workbook() -> Workbook:
    # Siempre libro nuevo: evita tablas huerfanas al recrear hojas.
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    crear_hoja_instrucciones(wb)
    crear_hoja_resultado(wb)
    crear_hoja_historial(wb)
    crear_hoja_captura(wb)
    if "Captura" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Captura")
    return wb


def main() -> None:
    wb = preparar_workbook()
    destino = SALIDA
    try:
        wb.save(destino)
    except PermissionError:
        destino = BASE_DIR / "Layout Comercial - Preparado (copia).xlsx"
        wb.save(destino)
        print("AVISO: cierre el Excel y vuelva a ejecutar para guardar en el archivo principal.")
    print("Listo:", destino)
    print("Hojas:", wb.sheetnames)


if __name__ == "__main__":
    main()
