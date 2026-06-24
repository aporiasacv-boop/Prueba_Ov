"""Prepara el Excel de marketing para pruebas con Dynamics."""
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

PANEL_FILAS = [
    ("U2", "Orden de venta (OV)", "V2", ""),
    ("U3", "Cuenta Dynamics", "V3", "C0010"),
    ("U4", "Nombre del cliente (opcional)", "V4", ""),
    ("U5", "Descripcion pedido", "V5", "Prueba desde excel odata"),
    ("U6", "Referencia (opcional)", "V6", ""),
    ("U7", "Fecha envio (opcional)", "V7", ""),
    ("U8", "Fecha recepcion (opcional)", "V8", ""),
]

PANEL_CAMPOS = {
    "ov": "V2",
    "cuenta": "V3",
    "nombre": "V4",
    "descripcion": "V5",
    "referencia": "V6",
    "fecha_envio": "V7",
    "fecha_recepcion": "V8",
}

BASE_DIR = Path(__file__).parent
CANDIDATOS = [
    BASE_DIR / "Prueba_OV marketing - Reparado.xlsx",
    BASE_DIR / "Prueba_OV marketing.xlsx",
]
SALIDA_FALLBACK = BASE_DIR / "Prueba_OV marketing - Preparado.xlsx"


def resolver_archivo() -> Path:
    for p in CANDIDATOS:
        if p.exists():
            return p
    return CANDIDATOS[-1]

COL_CUENTA = "Cuenta Dynamics"
COL_DESCRIPCION = "Descripción pedido"

RESULTADO_HEADERS = [
    "Estado",
    "Pedido Dynamics",
    "Fecha de ejecución",
    "Usuario",
    "Error",
]

HISTORIAL_HEADERS = [
    "FechaHora",
    "OrdenVenta",
    "ReferenciaCliente",
    "Cliente",
    "DescripcionPedido",
    "Codigo_Articulo",
    "Cantidad",
    "PrecioUnitario",
    "Porcentaje de descuento",
    "Fecha de envio",
    "Comentario",
]

INSTRUCCIONES = [
    ("A1", "SUBIR ORDENES DE VENTA A DYNAMICS (Marketing)"),
    ("A3", "1. Arranque dynamics-integration (run.ps1) y ngrok http 8080"),
    ("A4", "2. Pegue la URL https de ngrok en Resultado!B1"),
    ("A5", "3. Botones: ProbarConexion, CrearCabeceraMarketing, CrearLineasMarketing"),
    ("A7", "PANEL DYNAMICS (columnas U y V, filas 2-8 en Costco / E-commerce):"),
    ("A8", "- Etiquetas en columna U | Sus datos en columna V (al lado)"),
    ("A9", "- V2 = OV (se llena sola) | V3 = Cuenta (C0010) | V5 = Descripcion"),
    ("A10", "- Puede agregar filas a la TABLA de productos; el panel U:V no se mueve"),
    ("A12", "FLUJO: 1) Llene V3 y V5  2) Crear orden  3) Crear lineas"),
    ("A13", "Lineas: filas con Dynamics + Cantidad y sin OV en la tabla."),
]

TEST_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
TEST_FONT = Font(bold=True)

FILA_PRUEBA_ECOM = {
    "Fecha": "2026-06-09",
    "Mercado Libre": "Mercado Libre",
    "Fecha de compra": "2026-06-09",
    "N° venta": "TEST-OV-001",
    "Cliente": "Prueba OData (nombre)",
    COL_CUENTA: "C0010",
    COL_DESCRIPCION: "Prueba desde excel odata",
    "Dynamics": "503635001090",
    "Producto": "ISOFLAVONAS 90 TAB",
    "Cantidad": 1,
    "Costo ": 515.51,
    "Orden de venta": "",
}

FILA_PRUEBA_COSTCO = {
    "Fecha ": "2026-06-09",
    "Plataforma": "Costco",
    " PO": "TEST-COSTCO-001",
    "Cliente": "Prueba OData (nombre)",
    COL_CUENTA: "C0010",
    COL_DESCRIPCION: "Prueba desde excel odata",
    "Producto ": "JustFx 120 Capsulas",
    "Dynamics": "503635000705",
    "Cantidad": 1,
    "Ov": "",
}


def quitar_hoja_si_existe(wb, nombre: str) -> None:
    if nombre in wb.sheetnames:
        del wb[nombre]


def crear_hoja_resultado(wb) -> None:
    quitar_hoja_si_existe(wb, "Resultado")
    ws = wb.create_sheet("Resultado", 0)
    ws["A1"] = "URL API (ngrok)"
    ws["B1"] = ""
    ws["A2"] = "Resultado ultima ejecucion"
    for col, header in enumerate(RESULTADO_HEADERS, start=1):
        ws.cell(3, col, header)
    ws.cell(4, 1, "")

    tabla = Table(displayName="tblResultados", ref="A3:E4")
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["E"].width = 50


def crear_hoja_historial(wb) -> None:
    quitar_hoja_si_existe(wb, "Historial")
    ws = wb.create_sheet("Historial")
    for col, header in enumerate(HISTORIAL_HEADERS, start=1):
        ws.cell(1, col, header)

    last_col = get_column_letter(len(HISTORIAL_HEADERS))
    tabla = Table(displayName="tblHistorial", ref=f"A1:{last_col}1")
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)


def crear_hoja_instrucciones(wb) -> None:
    quitar_hoja_si_existe(wb, "Instrucciones Dynamics")
    ws = wb.create_sheet("Instrucciones Dynamics", 0)
    for addr, texto in INSTRUCCIONES:
        ws[addr] = texto
    ws.column_dimensions["A"].width = 95


def leer_encabezados(ws, tabla: Table) -> list[str]:
    min_col, min_row, max_col, _ = range_boundaries(tabla.ref)
    return [str(ws.cell(min_row, c).value or "").strip() for c in range(min_col, max_col + 1)]


def indice_encabezado(headers: list[str], nombre: str) -> int:
    objetivo = nombre.strip().lower()
    for i, h in enumerate(headers):
        if h.strip().lower() == objetivo:
            return i
    return -1


def agregar_columnas_si_faltan(ws, nombre_tabla: str, columnas: list[str]) -> None:
    tabla: Table = ws.tables[nombre_tabla]
    min_col, min_row, max_col, max_row = range_boundaries(tabla.ref)
    headers = leer_encabezados(ws, tabla)

    for nombre_col in columnas:
        if indice_encabezado(headers, nombre_col) >= 0:
            continue

        max_col += 1
        col_letra = get_column_letter(max_col)
        ws.cell(min_row, max_col, nombre_col)

        nueva = TableColumn(id=len(tabla.tableColumns) + 1, name=nombre_col)
        tabla.tableColumns.append(nueva)
        tabla.ref = f"{get_column_letter(min_col)}{min_row}:{col_letra}{max_row}"
        headers.append(nombre_col)


def fila_es_prueba(ws, tabla: Table, col_ov: str, valor_ref: str) -> int | None:
    min_col, min_row, max_col, max_row = range_boundaries(tabla.ref)
    headers = leer_encabezados(ws, tabla)
    idx_ref = indice_encabezado(headers, "N° venta")
    if idx_ref < 0:
        idx_ref = indice_encabezado(headers, " PO")
    if idx_ref < 0:
        return None

    for r in range(min_row + 1, max_row + 1):
        ref = str(ws.cell(r, min_col + idx_ref).value or "").strip()
        if ref == valor_ref:
            return r
    return None


def escribir_fila_por_headers(
    ws,
    tabla: Table,
    fila_num: int,
    valores: dict,
    resaltar: bool = False,
) -> None:
    min_col, min_row, _, _ = range_boundaries(tabla.ref)
    headers = leer_encabezados(ws, tabla)

    for nombre, valor in valores.items():
        idx = indice_encabezado(headers, nombre)
        if idx < 0:
            for h in headers:
                if h.strip().lower() == nombre.strip().lower():
                    idx = headers.index(h)
                    break
        if idx < 0:
            continue
        celda = ws.cell(fila_num, min_col + idx, valor)
        if resaltar:
            celda.fill = copy(TEST_FILL)
            celda.font = copy(TEST_FONT)


def insertar_fila_prueba(ws, nombre_tabla: str, valores: dict, col_ov: str, ref_prueba: str) -> None:
    tabla: Table = ws.tables[nombre_tabla]
    min_col, min_row, max_col, max_row = range_boundaries(tabla.ref)

    existente = fila_es_prueba(ws, tabla, col_ov, ref_prueba)
    if existente:
        escribir_fila_por_headers(ws, tabla, existente, valores, resaltar=True)
        return

    ws.insert_rows(min_row + 1)
    max_row += 1
    tabla.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    escribir_fila_por_headers(ws, tabla, min_row + 1, valores, resaltar=True)


def celda_absoluta(hoja: str, celda: str) -> str:
    col = "".join(ch for ch in celda if ch.isalpha())
    row = "".join(ch for ch in celda if ch.isdigit())
    return f"'{hoja}'!${col}${row}"


def definir_rangos_panel(wb, hoja: str, prefijo: str) -> None:
    for campo, celda in PANEL_CAMPOS.items():
        nombre = f"{prefijo}_{campo}"
        ref = celda_absoluta(hoja, celda)
        if nombre in wb.defined_names:
            del wb.defined_names[nombre]
        wb.defined_names.add(DefinedName(nombre, attr_text=ref))


def configurar_panel_dynamics(ws, prefijo: str) -> None:
    ws["U1"] = "PANEL DYNAMICS"
    ws["U1"].font = Font(bold=True)
    for celda_u, etiqueta, celda_v, valor_defecto in PANEL_FILAS:
        ws[celda_u] = etiqueta
        ws[celda_v] = valor_defecto
    definir_rangos_panel(ws.parent, ws.title, prefijo)


def preparar_tablas_captura(wb) -> None:
    configurar_panel_dynamics(wb["E-commerce"], "dyn_mkt_ecom")
    configurar_panel_dynamics(wb["Costco"], "dyn_mkt_costco")
    agregar_columnas_si_faltan(wb["E-commerce"], "Tabla4", [COL_CUENTA, COL_DESCRIPCION])
    agregar_columnas_si_faltan(wb["Costco"], "Tabla3", [COL_CUENTA, COL_DESCRIPCION])

    insertar_fila_prueba(
        wb["E-commerce"],
        "Tabla4",
        FILA_PRUEBA_ECOM,
        "Orden de venta",
        "TEST-OV-001",
    )
    insertar_fila_prueba(
        wb["Costco"],
        "Tabla3",
        FILA_PRUEBA_COSTCO,
        "Ov",
        "TEST-COSTCO-001",
    )


def main() -> None:
    origen = resolver_archivo()
    wb = load_workbook(origen)
    crear_hoja_instrucciones(wb)
    crear_hoja_resultado(wb)
    crear_hoja_historial(wb)
    preparar_tablas_captura(wb)

    destino = origen
    try:
        wb.save(destino)
    except PermissionError:
        destino = SALIDA_FALLBACK
        wb.save(destino)
        print("AVISO: cierre el Excel y vuelva a ejecutar para guardar en el archivo original.")
    print("Listo:", destino)
    print("Hojas:", wb.sheetnames[:6])


if __name__ == "__main__":
    main()
