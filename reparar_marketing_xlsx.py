"""Repara xlsx marketing: corrige XML y recrea Tabla3/Tabla4 validas."""
from __future__ import annotations

import shutil
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

BASE_DIR = Path(__file__).parent
ORIGEN = BASE_DIR / "Prueba_OV marketing.xlsx"
SALIDA = BASE_DIR / "Prueba_OV marketing - REPARADO.xlsx"

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
Q = f"{{{NS}}}"

TABLAS_CAPTURA = [
    ("Costco", "Tabla3", "N"),
    ("E-commerce", "Tabla4", "S"),
]

QUITAR_COLUMNAS = {"Cuenta Dynamics", "Descripción pedido", "Descripcion pedido"}


def qfind(el: ET.Element, tag: str) -> ET.Element | None:
    return el.find(Q + tag)


def limpiar_formula_calculada(col: ET.Element) -> bool:
    removio = False
    for hijo in list(col):
        if hijo.tag == Q + "calculatedColumnFormula":
            col.remove(hijo)
            removio = True
    return removio


def reparar_table_xml(xml: str, quitar_nombres: set[str]) -> tuple[str, list[str]]:
    cambios: list[str] = []
    root = ET.fromstring(xml)
    nombre = root.attrib.get("name", "?")
    cols_el = qfind(root, "tableColumns")
    if cols_el is None:
        return xml, cambios

    kept: list[ET.Element] = []
    for ch in list(cols_el):
        if ch.attrib.get("name", "") in quitar_nombres:
            cols_el.remove(ch)
            cambios.append(f"{nombre}: quitar columna tabla '{ch.attrib.get('name')}'")
        else:
            if limpiar_formula_calculada(ch):
                cambios.append(f"{nombre}: quitar formula calculada en '{ch.attrib.get('name')}'")
            kept.append(ch)

    for idx, ch in enumerate(kept, start=1):
        ch.set("id", str(idx))
    cols_el.set("count", str(len(kept)))

    ref = root.attrib.get("ref", "")
    if ref and kept:
        min_col, min_row, _, max_row = range_boundaries(ref)
        new_max_col = min_col + len(kept) - 1
        new_ref = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(new_max_col)}{max_row}"
        )
        if new_ref != ref:
            root.set("ref", new_ref)
            cambios.append(f"{nombre}: ref {ref} -> {new_ref}")
            ref = new_ref

    af = qfind(root, "autoFilter")
    if af is not None and ref:
        af.set("ref", ref)

    return ET.tostring(root, encoding="unicode"), cambios


def parchear_tablas_zip(destino: Path) -> list[str]:
    cambios: list[str] = []
    with zipfile.ZipFile(destino, "r") as zin:
        entries = list(zin.infolist())
        contenido = {info.filename: zin.read(info.filename) for info in entries}

    for archivo in ("xl/tables/table2.xml", "xl/tables/table3.xml"):
        if archivo not in contenido:
            continue
        xml = contenido[archivo].decode("utf-8")
        nuevo, delta = reparar_table_xml(xml, QUITAR_COLUMNAS)
        if nuevo != xml:
            contenido[archivo] = nuevo.encode("utf-8")
            cambios.extend(delta)

    with zipfile.ZipFile(destino, "w") as zout:
        for info in entries:
            zout.writestr(info, contenido[info.filename])
    return cambios


def ultima_fila_con_datos(ws, fila_inicio: int, col_inicio: int, col_fin: int) -> int:
    ultima = fila_inicio
    for fila in range(fila_inicio + 1, ws.max_row + 1):
        for col in range(col_inicio, col_fin + 1):
            if ws.cell(fila, col).value not in (None, ""):
                ultima = fila
                break
    return max(ultima, fila_inicio + 1)


def nombre_unico_tabla(headers: list[str], idx: int, nombre: str) -> str:
    base = (nombre or f"Columna{idx}").strip() or f"Columna{idx}"
    if base not in headers[: idx - 1]:
        return base
    suf = 2
    candidato = f"{base}_{suf}"
    while candidato in headers[: idx - 1]:
        suf += 1
        candidato = f"{base}_{suf}"
    return candidato


def recrear_tabla_captura(ws, nombre_tabla: str, col_max_letra: str) -> str:
    if nombre_tabla in ws.tables:
        del ws.tables[nombre_tabla]

    col_max = column_index_from_string(col_max_letra)
    fila_header = 1
    ultima = ultima_fila_con_datos(ws, fila_header, 1, col_max)

    headers: list[str] = []
    for col in range(1, col_max + 1):
        raw = ws.cell(fila_header, col).value
        headers.append(nombre_unico_tabla(headers, col, str(raw or "")))

    ref = f"A{fila_header}:{col_max_letra}{ultima}"
    tabla = Table(displayName=nombre_tabla, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight16",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    for i, header in enumerate(headers, start=1):
        tabla.tableColumns.append(TableColumn(id=i, name=header))

    ws.add_table(tabla)
    return ref


def reparar_workbook(destino: Path) -> list[str]:
    cambios: list[str] = []
    wb = load_workbook(destino)
    for hoja, nombre_tabla, col_max in TABLAS_CAPTURA:
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        ref = recrear_tabla_captura(ws, nombre_tabla, col_max)
        cambios.append(f"{hoja}/{nombre_tabla}: recreada en {ref}")
    wb.save(destino)
    return cambios


def reparar_xlsx(origen: Path, destino: Path) -> list[str]:
    shutil.copy2(origen, destino)
    cambios = parchear_tablas_zip(destino)
    cambios.extend(reparar_workbook(destino))
    return cambios


def main() -> None:
    origen = Path(sys.argv[1]) if len(sys.argv) > 1 else ORIGEN
    destino = Path(sys.argv[2]) if len(sys.argv) > 2 else SALIDA
    if not origen.exists():
        raise FileNotFoundError(
            f"No se encontro {origen}. Restaure el Excel original desde OneDrive."
        )

    cambios = reparar_xlsx(origen, destino)
    print("Listo:", destino)
    for linea in cambios:
        print(" -", linea)


if __name__ == "__main__":
    main()
