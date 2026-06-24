"""Diagnostico profundo de tablas marketing vs hoja Excel."""
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
Q = f"{{{NS}}}"


def headers_hoja(path: Path, sheet: str, fila: int, max_col: int) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    out = []
    for c in range(1, max_col + 1):
        v = ws.cell(fila, c).value
        out.append("" if v is None else str(v).strip())
    wb.close()
    return out


def columnas_tabla_xml(path: Path, table_file: str) -> list[tuple[str, str]]:
    with zipfile.ZipFile(path) as z:
        xml = z.read(table_file).decode("utf-8", errors="replace")
    root = ET.fromstring(xml)
    ref = root.attrib.get("ref", "")
    cols = []
    cols_el = root.find(Q + "tableColumns")
    if cols_el is not None:
        for ch in cols_el:
            cols.append((ch.attrib.get("id", "?"), ch.attrib.get("name", "?")))
    return cols, ref, xml


def diagnostico(path: Path) -> None:
    print("\n==========", path.name, "==========")
    mapping = {
        "xl/tables/table2.xml": ("Costco", "Tabla3"),
        "xl/tables/table3.xml": ("E-commerce", "Tabla4"),
    }
    for tf, (sheet, tname) in mapping.items():
        if not path.exists():
            continue
        try:
            cols, ref, _ = columnas_tabla_xml(path, tf)
        except KeyError:
            print(tf, "missing")
            continue
        print(f"\n{tf} ({tname}) ref={ref}")
        ids = [c[0] for c in cols]
        names = [c[1] for c in cols]
        print("  xml ids:", ids)
        print("  xml names:", names)
        if ids != [str(i) for i in range(1, len(ids) + 1)]:
            print("  !! ids no secuenciales")

        min_col, min_row, max_col, max_row = range_boundaries(ref)
        hoja = headers_hoja(path, sheet, min_row, max_col)
        print("  hoja headers:", hoja)
        span = max_col - min_col + 1
        if len(names) != span:
            print("  !! count mismatch xml vs ref span")
        mism = []
        for i in range(min(len(names), len(hoja))):
            if names[i].strip() != hoja[i].strip() and names[i].strip().lower() != hoja[i].strip().lower():
                mism.append((i + 1, names[i], hoja[i]))
        if mism:
            print("  !! header mismatches (pos, xml, hoja):")
            for m in mism[:8]:
                print("     ", m)
            if len(mism) > 8:
                print("      ...", len(mism) - 8, "more")


def test_openpyxl_roundtrip(origen: Path) -> None:
    dest = origen.parent / (origen.stem + "-ROUNDTRIP.xlsx")
    wb = load_workbook(origen)
    wb.save(dest)
    print("\n--- despues de openpyxl save sin cambios ---")
    diagnostico(dest)


def main() -> None:
    base = Path(__file__).parent
    files = [
        base / "Prueba_OV marketing.xlsx",
        base / "Prueba_OV marketing - REPARADO.xlsx",
        base / "Prueba_OV marketing - Preparado.xlsx",
    ]
    for f in files:
        if f.exists():
            diagnostico(f)
    rep = base / "Prueba_OV marketing - REPARADO.xlsx"
    if rep.exists():
        test_openpyxl_roundtrip(rep)

if __name__ == "__main__":
    main()
